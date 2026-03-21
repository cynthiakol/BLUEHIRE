import random
import uuid
import datetime
import openpyxl
import traceback
import os
import pytesseract
import PyPDF2

from systemlogs.models import AIDocumentCheck
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.views import PasswordChangeView
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.conf import settings
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth import logout, authenticate, login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_POST
from django.contrib.auth.hashers import make_password
from django.core.mail import send_mail
from django.conf import settings
from django.db.models import Q, Count
from django.urls import reverse
from django.middleware.csrf import get_token
from django.shortcuts import get_object_or_404
from accounts.models import PhoneOTP
from core.config import BLUEHIRE_VERSION
from accounts.models import JobSeeker
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from jobs.matcher import compute_similarity
from jobs.matcher import recommend_jobs_for_seeker
from accounts.doc_verifier import check_document_validity
from accounts.models import DocumentVerification
from PIL import Image
from django.core.files.storage import default_storage
# ==================== MODELS ====================
from accounts.models import (
    Users,
    JobSeeker,
    Employer,
    PhoneOTP,
    SystemLog,
    EmployerDocuments,
    JobSeekerCertificate,
)
from jobs.models import Job
from applications.models import Application
from jobs.models import Notification
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_POST
from accounts.doc_verifier import run_ai_verification_for_user
from django.db import models
import json
from jobs.ai_matcher import rank_applicants
from django.db.models import Avg
from ratings.models import Rating

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# ====================================================
# 🔹 SEND OTP (AJAX)
# ====================================================
import smtplib
from email.mime.text import MIMEText
from django.conf import settings

@csrf_exempt
def send_otp(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid method"}, status=405)

    try:
        data = json.loads(request.body.decode("utf-8"))
        phone = data.get("phone_number")
        carrier = data.get("carrier", "globe")
        
        if not phone:
            return JsonResponse({"status": "error", "message": "Phone number required"})

        # Generate OTP
        otp_code = random.randint(100000, 999999)
        PhoneOTP.objects.update_or_create(
            phone=phone,
            defaults={"otp": otp_code, "created_at": timezone.now()}
        )

        print(f"[DEBUG] ✅ OTP generated for {phone}: {otp_code}")

        # ✅ TRY TO SEND SMS (but don't fail if it doesn't work)
        sms_sent = False
        try:
            carrier_gateways = {
                'globe': '@email2sms.globe.com.ph',
                'smart': '@sms.smart.com.ph',
                'sun': '@sun.com.ph',
                'tnt': '@sms.smart.com.ph',
            }
            
            # Clean phone number
            clean_phone = phone.replace('+63', '').replace(' ', '').strip()
            if clean_phone.startswith('0'):
                clean_phone = clean_phone[1:]
            
            # Create SMS email
            to_email = clean_phone + carrier_gateways.get(carrier, carrier_gateways['globe'])
            
            msg = MIMEText(f"Your BlueHire OTP is: {otp_code}")
            msg['From'] = settings.EMAIL_HOST_USER
            msg['To'] = to_email
            msg['Subject'] = "BlueHire OTP"
            
            # Send via Gmail
            with smtplib.SMTP(settings.EMAIL_HOST, settings.EMAIL_PORT) as server:
                server.starttls()
                server.login(settings.EMAIL_HOST_USER, settings.EMAIL_HOST_PASSWORD)
                server.send_message(msg)
            
            sms_sent = True
            print(f"[DEBUG] ✅ SMS sent to {to_email}")
            
        except Exception as sms_error:
            print(f"[WARNING] SMS sending failed: {sms_error}")
            # Don't fail the whole request - just log the error

        # ✅ ALWAYS RETURN SUCCESS (OTP is saved in DB)
        return JsonResponse({
            "status": "ok",
            "message": "OTP generated successfully",
            "otp": otp_code,  # For testing - remove in production
            "sms_sent": sms_sent
        })

    except Exception as e:
        print(f"[ERROR] send_otp failed: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({"status": "error", "message": str(e)}, status=500)
# ====================================================
# 🔹 VERIFY OTP (AJAX)
# ====================================================
@csrf_exempt
def verify_otp(request):
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid method"}, status=405)

    try:
        phone = request.POST.get("phone") or request.POST.get("phone_number")
        otp = request.POST.get("otp")

        if not phone or not otp:
            return JsonResponse({"status": "error", "message": "Missing phone or OTP"})

        # Get OTP record
        record = PhoneOTP.objects.filter(phone=phone).first()

        if not record:
            return JsonResponse({"status": "error", "message": "No OTP found for this number"})

        # ✅ Compare OTP safely
        if str(record.otp).strip() == str(otp).strip():
            # ✅ Delete after successful verification (one-time use)
            record.delete()
            print(f"[DEBUG] ✅ OTP verified for {phone}")
            return JsonResponse({
                "status": "ok",
                "message": "OTP verified successfully!"
            })
        else:
            print(f"[DEBUG] ❌ Invalid OTP for {phone}. Expected: {record.otp}, Got: {otp}")
            return JsonResponse({
                "status": "error",
                "message": "Invalid OTP. Please check and try again."
            })

    except Exception as e:
        print(f"[ERROR] verify_otp failed: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({"status": "error", "message": f"Server error: {e}"}, status=500)

# ====================================================
# 🔹 EMAIL VERIFICATION UTIL
# ====================================================
def send_verification_email(user):
    """Send verification email with unique token"""
    if not user.email_token:
        user.email_token = uuid.uuid4()
        user.save()

    verify_link = f"{settings.SITE_URL}{reverse('verify_email_token', args=[user.email_token])}"
    subject = "Verify your email - BlueHire"
    message = (
        f"Hi {user.username or 'User'},\n\n"
        f"Please verify your email by clicking the link below:\n"
        f"{verify_link}\n\n"
        f"If you didn’t create this account, you can safely ignore this email.\n\n"
        "Thank you for joining BlueHire!"
    )

    try:
        send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [user.email])
        print(f"[BlueHire] verification email sent to {user.email}")
    except Exception as e:
        print(f"Error sending verification email: {e}")

# ====================================================
# 🔹 SIGNIN + REGISTER
# ====================================================
def signin_register(request):
    # ✅ NO CHECK FOR AUTHENTICATED USERS HERE
    
    if request.method == "POST":
        action = request.POST.get("action")

        # ====================================================
        # 🔹 REGISTER
        # ====================================================
        if action == "register":
            username = request.POST.get("username", "").strip()
            first_name = request.POST.get("first_name", "").strip()
            last_name = request.POST.get("last_name", "").strip()
            email = request.POST.get("email", "").strip()
            phone = request.POST.get("phone_number", "").strip()
            password = request.POST.get("password", "")
            role = request.POST.get("role", "")
            employer_type = request.POST.get("employer_type", "")

            # Basic validation
            if Users.objects.filter(email=email).exists():
                messages.error(request, "Email already used.")
                return redirect("signin_register")

            if Users.objects.filter(username=username).exists():
                messages.error(request, "Username already taken.")
                return redirect("signin_register")

            # ✅ Validate phone number
            if not phone:
                messages.error(request, "Phone number is required.")
                return redirect("signin_register")

            # Create user
            new_user = Users.objects.create(
                username=username,
                first_name=first_name,
                last_name=last_name,
                email=email,
                phone=phone,
                password=make_password(password),
                role=role,
                is_verified=False,
                email_token=uuid.uuid4(),
            )

            # ✅ Debug print to verify phone is saved
            print(f"[DEBUG] User created: {new_user.username}, Phone: {new_user.phone}")

            # ✅ CREATE RELATED RECORD WITH "INCOMPLETE" STATUS
            if role == "JobSeeker":
                seeker = JobSeeker.objects.create(user=new_user)
                seeker.approval_status = "Incomplete"
                seeker.verification_status = "Incomplete"
                seeker.save()
                
            elif role == "Employer":
                employer = Employer.objects.create(user=new_user, employer_type=employer_type)
                employer.approval_status = "Incomplete"
                employer.save()

            # Send verification email
            send_verification_email(new_user)
            
            # ✅ Authenticate the user FIRST to set the backend attribute
            authenticated_user = authenticate(request, username=username, password=password)
            
            if authenticated_user:
                # Log the user in
                login(request, authenticated_user)
                messages.success(request, "Account created successfully! Please verify your email.")
                
                # ✅ Redirect based on role
                if role == "JobSeeker":
                    return redirect("jobseeker_dashboard")
                elif role == "Employer":
                    return redirect("employer_dashboard")
                elif role == "Admin":
                    return redirect("system_admin_dashboard")
                else:
                    return redirect("homepage")
            else:
                # If authentication fails (shouldn't happen), set backend manually
                new_user.backend = 'django.contrib.auth.backends.ModelBackend'
                login(request, new_user)
                messages.success(request, "Account created successfully! Please verify your email.")
                
                # ✅ Redirect based on role
                if role == "JobSeeker":
                    return redirect("jobseeker_dashboard")
                elif role == "Employer":
                    return redirect("employer_dashboard")
                elif role == "Admin":
                    return redirect("system_admin_dashboard")
                else:
                    return redirect("homepage")

        # ====================================================
        # 🔹 SIGN IN
        # ====================================================
        elif action == "signin":
            username = request.POST.get("username")
            password = request.POST.get("password")

            user = authenticate(request, username=username, password=password)

            if user is not None:
                login(request, user)
                request.session.set_expiry(0)

                if not user.is_verified and user.role != "Admin":
                    messages.warning(
                        request,
                        "📩 Your email is not verified. Please check your inbox or click 'Verify now' on your profile."
                    )

                # role-based redirects
                if user.role == "JobSeeker":
                    return redirect("jobseeker_dashboard")

                elif user.role == "Employer":
                    return redirect("employer_dashboard")

                elif user.role == "Admin" or user.is_superuser:
                    return redirect("system_admin_dashboard")

                else:
                    return redirect("homepage")

            messages.error(request, "Invalid username or password.")
            return redirect("signin_register")

    # ====================================================
    # 🔹 GET Request (default render)
    # ====================================================
    return render(request, "core/signin_register.html")


# ====================================================
# 🔹 EMAIL VERIFICATION FLOW
# ====================================================

from django.utils import timezone
import uuid
from django.core.mail import send_mail
from django.contrib import messages
from django.conf import settings
from django.contrib.auth.decorators import login_required
from accounts.models import Users
from jobs.models import Notification


@login_required
def verify_email(request):
    """Send verification link to the user’s email"""
    user = request.user
    if not user.is_verified:
        token = uuid.uuid4()
        user.email_token = token
        user.save()

        verify_link = f"{settings.SITE_URL}/accounts/verify-email/{token}/"
        send_mail(
            subject="Verify your BlueHire Email",
            message=(
                f"Hey {user.first_name},\n\n"
                f"Click this link to verify your BlueHire account:\n\n{verify_link}\n\n"
                f"If you see this email, it means our verification system is working perfectly 🎉"
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
        )

        # ✅ this shows on your profile after clicking "Verify Email"
        messages.success(request, "Verification link sent! Please check your inbox.")
    else:
        messages.info(request, "Your email is already verified.")

    return redirect("jobseeker_profile")


def verify_email_token(request, token):
    try:
        user = Users.objects.get(email_token=token)
        user.is_verified = True
        user.email_verified_at = timezone.now()
        user.email_token = None  # ✅ correct field reset
        user.save()

        # ✅ notification creation using correct field name
        Notification.objects.create(
            receiver=user,  # use receiver instead of user since your Notification model uses 'receiver'
            message="Your email has been successfully verified."
        )

        messages.success(request, "Your email has been successfully verified.")
        return redirect("jobseeker_profile")

    except Users.DoesNotExist:
        messages.error(request, "Invalid or expired verification link.")
        return redirect("signin_register")

@login_required
def resend_verification(request):
    user = request.user
    if user.is_verified:
        messages.info(request, "Your email is already verified.")
    else:
        send_verification_email(user)  # ✅ keep your email-sending function
        messages.success(request, "Verification link sent to your email.")

    # 🩵 stay on the same profile page depending on user type
    if hasattr(user, "employer"):
        return redirect("employer_profile")
    elif hasattr(user, "jobseeker"):
        return redirect("jobseeker_profile")
    else:
        return redirect("home")


# ====================================================
# 🔹 PROFILE VIEWS
# ====================================================
@login_required
def profile_view(request):
    user = request.user
    seeker = JobSeeker.objects.filter(user=user).first()
    employer = Employer.objects.filter(user=user).first()

    return render(
        request,
        "core/jobseeker_profile.html",
        {"user": user, "seeker": seeker, "employer": employer},
    )

@login_required
def jobseeker_profile(request):
    seeker = JobSeeker.objects.filter(user=request.user).first()
    if not seeker:
        messages.error(request, "Jobseeker profile not found.")
        return redirect("jobseeker_dashboard")

    # Handle updates (this is handled by update_jobseeker_profile now)
    if request.method == "POST":
        return redirect("update_jobseeker_profile")

    # Notifications
    notifications = Notification.objects.filter(receiver=request.user).order_by("-created_at")[:10]
    unread_qs = Notification.objects.filter(receiver=request.user, is_read=False)
    unread_count = unread_qs.count()
    unread_qs.update(is_read=True)

    # Fetch verification records with CORRECT status - ✅ NOW INCLUDES BACK SIDE
    verifications = {
        "Police_Clearance": DocumentVerification.objects.filter(
            user=request.user, 
            doc_type="Police Clearance"
        ).first(),
        "Nbi_Clearance": DocumentVerification.objects.filter(
            user=request.user, 
            doc_type="Nbi Clearance"
        ).first(),
        "Valid_Id": DocumentVerification.objects.filter(
            user=request.user, 
            doc_type="Valid Id"
        ).first(),
        "Valid_Id_Back": DocumentVerification.objects.filter(
            user=request.user, 
            doc_type="Valid Id Back"
        ).first(),
        "Barangay_Clearance": DocumentVerification.objects.filter(
            user=request.user, 
            doc_type="Barangay Clearance"
        ).first(),
    }

    # Job type options
    job_types = [
        "Construction", "Manufacturing", "Maintenance", "Transportation",
        "Plumber", "Customer Service", "Carpentry", "Driver",
        "Technician", "Electrician", "Mechanic", "Laborer", "Other"
    ]

    # ✅ UPDATED: Document fields list - NOW SAFE
    document_fields = [
        ("Police Clearance", seeker.police_clearance, verifications.get("Police_Clearance")),
        ("Nbi Clearance", seeker.nbi_clearance, verifications.get("Nbi_Clearance")),
        ("Valid Id", seeker.valid_id, verifications.get("Valid_Id")),
        ("Valid Id Back", seeker.valid_id_back, verifications.get("Valid_Id_Back")),  # ✅ NOW SAFE
        ("Barangay Clearance", seeker.barangay_clearance, verifications.get("Barangay_Clearance")),
    ]

    seeker.refresh_from_db()

    context = {
        "seeker": seeker,
        "job_types": job_types,
        "notifications": notifications,
        "unread_count": unread_count,
        "verifications": verifications,
        "document_fields": document_fields,
    }

    return render(request, "core/jobseeker_profile.html", context)

@login_required
def update_jobseeker_profile(request):
    """Handle jobseeker profile updates including NAME and PHONE"""
    seeker = JobSeeker.objects.filter(user=request.user).first()
    if not seeker:
        messages.error(request, "Jobseeker profile not found.")
        return redirect("jobseeker_dashboard")

    if request.method == "POST":
        # Handle profile picture upload
        if "profile_photo" in request.FILES:
            seeker.profile_photo = request.FILES["profile_photo"]
            seeker.save()
            messages.success(request, "Profile picture updated successfully!")
            return redirect("jobseeker_profile")
        
        # ✅ NEW: Update User model fields (first_name, last_name, phone)
        user = request.user
        user.first_name = request.POST.get("first_name", "").strip()
        user.last_name = request.POST.get("last_name", "").strip()
        user.phone = request.POST.get("phone", "").strip()
        user.save()
        
        # Handle other profile updates (JobSeeker model)
        seeker.gender = request.POST.get("gender", "")
        seeker.birthdate = request.POST.get("birthdate") or None
        seeker.region = request.POST.get("region", "")
        seeker.province = request.POST.get("province", "")
        seeker.city = request.POST.get("city", "")
        seeker.barangay = request.POST.get("barangay", "")
        seeker.about = request.POST.get("about", "")
        seeker.contact_number = request.POST.get("contact_number", "")

        preferred_job_types = request.POST.getlist("preferred_job_types")
        seeker.preferred_job_types = preferred_job_types if preferred_job_types else []

        seeker.save()

        # Recalculate profile completion
        fields = [
            seeker.gender,
            seeker.birthdate,
            seeker.region,
            seeker.city,
            seeker.about,
            seeker.profile_photo,
            seeker.contact_number,
            seeker.police_clearance,
            seeker.nbi_clearance,
            seeker.valid_id,
            user.first_name,  # ✅ Include name in completion
            user.last_name,
        ]
        completed = sum(1 for f in fields if f)
        seeker.profile_completion = int((completed / len(fields)) * 100)
        seeker.save(update_fields=["profile_completion"])

        messages.success(request, "Profile updated successfully!")
        return redirect("jobseeker_profile")

    return redirect("jobseeker_profile")


@login_required
def upload_jobseeker_documents(request):
    seeker = JobSeeker.objects.filter(user=request.user).first()
    if not seeker:
        messages.error(request, "JobSeeker profile not found.")
        return redirect("jobseeker_dashboard")

    # Block rejected users from uploading
    if seeker.approval_status == "Rejected":
        messages.error(request, "Your account has been rejected. You can no longer upload or edit your documents.")
        return redirect("jobseeker_dashboard")

    if request.method == "POST":
        # Updated files map - NOW INCLUDES BACK SIDE
        files_map = {
            "Barangay Clearance": request.FILES.get("barangay-clearance"),
            "Police Clearance": request.FILES.get("police-clearance"),
            "Nbi Clearance": request.FILES.get("nbi-clearance"),
            "Valid Id": request.FILES.get("valid-id"),
            "Valid Id Back": request.FILES.get("valid-id-back"),  # ✅ ADDED
        }

        valid_id_type = request.POST.get("valid_id_type")

        from jobs.models import Notification
        from accounts.models import DocumentVerification
        from systemlogs.models import AIDocumentCheck
        from PIL import Image
        import pytesseract

        pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

        for field_name, file in files_map.items():
            if not file:
                continue

            # Map field name to model field
            if field_name == "Valid Id":
                model_field = "valid_id"
            elif field_name == "Valid Id Back":
                model_field = "valid_id_back"  # ✅ ADDED
            else:
                model_field = field_name.lower().replace(" ", "_")
            
            # Save the file to the model
            setattr(seeker, model_field, file)
            if field_name == "Valid Id" and valid_id_type:
                seeker.valid_id_type = valid_id_type
            
            # IMPORTANT: Save after setting each file
            seeker.save()

            # OCR AI scan
            try:
                file_path = getattr(seeker, model_field).path
                with Image.open(file_path) as img:
                    ocr = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
                    text = " ".join(ocr.get("text", [])).strip()
                    confs = [float(c) for c in ocr.get("conf", []) if c not in ["-1", ""]]
                    conf = (sum(confs) / len(confs) / 100) if confs else 0.0
            except Exception as e:
                text = f"OCR error: {e}"
                conf = 0.0

            # Create AI log
            AIDocumentCheck.objects.create(
                jobseeker=seeker,
                document_type=field_name,
                confidence_score=conf,
                detected_text=text,
            )

            # Threshold-based document flagging
            if conf < 0.40:
                doc_status = "Re-upload Required"
                Notification.objects.create(
                    receiver=request.user,
                    message=f"Your {field_name} appears blurry or unclear. Please re-upload a clearer version."
                )
            else:
                doc_status = "Pending"

            # Create/update verification record
            DocumentVerification.objects.update_or_create(
                user=request.user,
                doc_type=field_name,
                defaults={
                    "file": getattr(seeker, model_field),
                    "status": doc_status,
                    "ai_confidence": conf,
                },
            )

        # ✅ CHECK IF ALL REQUIRED DOCUMENTS ARE UPLOADED (INCLUDING BACK SIDE)
        all_docs_uploaded = all([
            seeker.barangay_clearance,
            seeker.police_clearance,
            seeker.nbi_clearance,
            seeker.valid_id,
            seeker.valid_id_back  # ✅ ADDED
        ])

        # ✅ ONLY SET TO PENDING IF ALL DOCS ARE UPLOADED
        if all_docs_uploaded:
            seeker.approval_status = "Pending"
            seeker.verification_status = "Pending"
            
            # ✅ Notify admins that documents are ready for review
            from django.contrib.auth import get_user_model
            User = get_user_model()
            admins = User.objects.filter(role="Admin")
            for admin in admins:
                Notification.objects.create(
                    receiver=admin,
                    message=f"{request.user.first_name} {request.user.last_name} has uploaded all documents and is ready for verification.",
                    created_at=timezone.now(),
                )
            
            messages.success(request, "All documents uploaded! Your application is now pending admin review.")
        else:
            seeker.approval_status = "Incomplete"
            seeker.verification_status = "Incomplete"
            messages.success(request, "Documents uploaded. Please upload all required documents to proceed.")
        
        seeker.save()

        return redirect("jobseeker_profile")

    messages.error(request, "Invalid request.")
    return redirect("jobseeker_profile")



# ====================================================
# 🔹 JOB SEEKER DOCUMENTS
# ====================================================
@login_required
def upload_jobseeker_certificates(request):
    seeker = JobSeeker.objects.filter(user=request.user).first()  # ✅ FIXED
    if not seeker:
        messages.error(request, "JobSeeker profile not found.")
        return redirect("jobseeker_dashboard")

    if request.method == "POST":
        cert_name = request.POST.get("certificate_name")
        cert_file = request.FILES.get("certificate_file")

        if not cert_name or not cert_file:
            messages.error(request, "Please enter a certificate name and choose a file.")
            return redirect("jobseeker_profile")

        JobSeekerCertificate.objects.create(
            jobseeker=seeker,
            name=cert_name,
            file=cert_file
        )
        messages.success(request, f"Certificate '{cert_name}' uploaded successfully.")
        return redirect("jobseeker_profile")

    return redirect("jobseeker_profile")

def update_jobseeker_progress(seeker):
    """Enhanced profile completion formula that includes personal info, experience, and documents."""
    completion = 0
    total_fields = 12  # we’ll grade on 12 key items for better balance

    # 🧍 personal details
    if seeker.profile_photo:
        completion += 1
    if seeker.gender:
        completion += 1
    if getattr(seeker, "birthdate", None):
        completion += 1
    if getattr(seeker, "region", None) and getattr(seeker, "city", None):
        completion += 1
    if getattr(seeker, "about", None):
        completion += 1
    if getattr(seeker, "contact_number", None):
        completion += 1

    # 💼 experience & skills
    if getattr(seeker, "skills", None):
        completion += 1
    if getattr(seeker, "experience", None):
        completion += 1
    if getattr(seeker, "education", None):
        completion += 1

    # 📄 documents (each adds small credit)
    if seeker.barangay_clearance:
        completion += 1
    if seeker.police_clearance:
        completion += 1
    if seeker.nbi_clearance:
        completion += 1
    if seeker.valid_id:
        completion += 1

    # ✅ compute percentage safely
    seeker.profile_completion = int((completion / total_fields) * 100)
    seeker.save(update_fields=["profile_completion"])
    return seeker.profile_completion


# ====================================================
# 🔹 JOB SEEKER DASHBOARD
# ====================================================
# Replace the jobseeker_dashboard function in accounts/views.py

@login_required
def jobseeker_dashboard(request):
    user = request.user
    seeker = JobSeeker.objects.filter(user=user).first()
    employers = Employer.objects.filter(employer_type="Company")

    # ✅ Display user name safely
    user_name = (
        f"{getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')}".strip()
        or getattr(user, "name", None)
        or getattr(user, "username", None)
        or (user.email.split('@')[0] if user.email else "User")
    )

    # ✅ Filters for search
    q = request.GET.get("q", "")
    job_type = request.GET.get("category", "")
    company = request.GET.get("company", "")

    # ✅ Get jobs where user was hired - EXCLUDE THESE
    hired_job_ids = []
    if seeker:
        hired_job_ids = Application.objects.filter(
            applicant=seeker,
            status="Hired"
        ).values_list('job_id', flat=True)

    # ✅ Filter jobs (only approved) with STATS ANNOTATION
    from django.db.models import Count, Q
    
    jobs = Job.objects.filter(is_approved=True).select_related(
        'employer', 'employer__user'
    ).order_by("-created_at")
    
    # ✅ Exclude hired jobs
    if hired_job_ids:
        jobs = jobs.exclude(id__in=hired_job_ids)

    if q:
        jobs = jobs.filter(
            Q(title__icontains=q)
            | Q(description__icontains=q)
            | Q(employer__company_name__icontains=q)
        )
    if job_type:
        jobs = jobs.filter(category__icontains=job_type)
    if company:
        jobs = jobs.filter(employer__company_name__icontains=company)

    # ✅ MANUALLY CALCULATE STATS FOR EACH JOB (FIXES 0 APPLICANTS BUG)
    jobs_list = list(jobs)
    for job in jobs_list:
        # Count all applications for this job
        job.total_applicants = Application.objects.filter(job=job).count()
        job.hired_count = Application.objects.filter(job=job, status='Hired').count()
        job.rejected_count = Application.objects.filter(job=job, status='Rejected').count()
        job.pending_count = Application.objects.filter(job=job, status='Pending').count()

    # ✅ AI-powered job recommendations
    recommended_jobs = []
    if seeker and seeker.skills:
        try:
            available_for_recommendation = Job.objects.filter(
                is_approved=True
            ).select_related('employer', 'employer__user')
            
            if hired_job_ids:
                available_for_recommendation = available_for_recommendation.exclude(id__in=hired_job_ids)
            
            # Calculate stats for recommended jobs too
            available_list = list(available_for_recommendation)
            for job in available_list:
                job.total_applicants = Application.objects.filter(job=job).count()
                job.hired_count = Application.objects.filter(job=job, status='Hired').count()
                job.rejected_count = Application.objects.filter(job=job, status='Rejected').count()
            
            ranked = recommend_jobs_for_seeker(seeker, available_list)
            recommended_jobs = [job for job, _ in ranked[:5]]
        except Exception as e:
            print("AI recommendation error:", e)
            recommended_jobs = jobs_list[:5]
    else:
        recommended_jobs = jobs_list[:5]

    # ✅ Application tracking (status + recent) WITH STATS
    applications = Application.objects.filter(
        applicant=seeker
    ).select_related(
        "job", 
        "job__employer",
        "job__employer__user"
    ) if seeker else []
    
    applied_job_ids = [app.job_id for app in applications] if applications else []
    applied_job_statuses = {app.job_id: app.status for app in applications} if applications else {}

    # ✅ Recently applied jobs (latest 5) - MANUALLY ADD STATS
    recent_applications = []
    if applications:
        recent_apps = applications.order_by("-applied_date")[:5]
        for app in recent_apps:
            # Manually add stats to each job
            job = app.job
            job.total_applicants = Application.objects.filter(job=job).count()
            job.hired_count = Application.objects.filter(job=job, status='Hired').count()
            job.rejected_count = Application.objects.filter(job=job, status='Rejected').count()
            recent_applications.append(app)

    # ✅ Categories
    categories = Job.objects.values_list("category", flat=True).distinct()

    # ✅ Document verification check
    missing_docs = not (
        seeker and seeker.barangay_clearance and seeker.nbi_clearance and seeker.police_clearance
    )

    # ✅ Profile completion
    doc_fields = [seeker.barangay_clearance, seeker.nbi_clearance, seeker.police_clearance] if seeker else []
    filled = sum(1 for f in doc_fields if f)
    total = len(doc_fields) if doc_fields else 1
    progress_percent = int((filled / total) * 100)

    context = {
        "user_name": user_name,
        "seeker": seeker,
        "jobs": jobs_list,  # ✅ Pass list with calculated stats
        "recommended_jobs": recommended_jobs,
        "employers": employers,
        "applied_job_ids": applied_job_ids,
        "applied_job_statuses": applied_job_statuses,
        "recent_applications": recent_applications,  # ✅ Now has stats
        "categories": categories,
        "missing_docs": missing_docs,
        "progress_percent": progress_percent,
    }

    return render(request, "core/jobseeker_dashboard.html", context)


@login_required
def jobseeker_public_profile(request, jobseeker_id):
    from ratings.models import Rating
    from django.db.models import Count

    jobseeker = get_object_or_404(JobSeeker, pk=jobseeker_id)

    location_parts = [
        jobseeker.barangay or "",
        jobseeker.city or "",
        jobseeker.province or "",
        jobseeker.region or "",
    ]
    location = ", ".join([part for part in location_parts if part])

    preferred_job_types = jobseeker.preferred_job_types or []
    skills = jobseeker.skills or []
    certifications = jobseeker.certifications or []

    clearances = {
        "Barangay Clearance": jobseeker.barangay_clearance,
        "Police Clearance": jobseeker.police_clearance,
        "NBI Clearance": jobseeker.nbi_clearance,
    }

    job_types = [
        "Construction", "Manufacturing", "Maintenance", "Transportation",
        "Plumber", "Customer Service", "Carpentry", "Driver",
        "Technician", "Electrician", "Mechanic", "Laborer", "Other"
    ]

    ratings = Rating.objects.filter(
        rated_jobseeker=jobseeker, is_visible=True
    ).select_related('reviewer', 'application__job')
    stats = ratings.aggregate(avg=Avg('score'), total=Count('id'))
    avg_score = round(stats['avg'] or 0, 1)
    avg_rating = avg_score if stats['total'] > 0 else None
    total_ratings = stats['total']
    distribution = {i: ratings.filter(score=i).count() for i in range(5, 0, -1)}

    return render(
        request,
        "core/jobseeker_public_profile.html",
        {
            "jobseeker": jobseeker,
            "job_types": job_types,
            "location": location or "N/A",
            "preferred_job_types": preferred_job_types,
            "skills": skills,
            "certifications": certifications,
            "clearances": clearances,
            "ratings": ratings,
            "avg_rating": avg_rating,
            "total_ratings": total_ratings,
            "distribution": distribution,
            "filled_stars": range(1, round(avg_score) + 1),
            "empty_stars": range(round(avg_score) + 1, 6),
        },
    )

@login_required
def update_skills(request):
    seeker = JobSeeker.objects.filter(user=request.user).first()  # ✅ FIXED
    if request.method == "POST" and seeker:
        seeker.skills = [s.strip() for s in request.POST.get("skills","").split(",") if s.strip()]
        seeker.preferred_job_types = request.POST.getlist("preferred_job_types")
        seeker.save()
        messages.success(request, "Updated skills and preferred job types.")
    return redirect("jobseeker_profile")
# ====================================================
# 🔹 APPLY FOR JOB
# ====================================================


@login_required
def apply_job(request, job_id):
    seeker = JobSeeker.objects.filter(user=request.user).first()

    if not seeker:
        messages.error(request, "Jobseeker profile not found.")
        return redirect("jobseeker_dashboard")

    if seeker.approval_status != "Approved":
        messages.warning(request, "Your account is pending admin approval. You'll be notified once reviewed.")
        return redirect("jobseeker_dashboard")

    job = get_object_or_404(Job, id=job_id, is_available=True, is_approved=True)

    # prevent duplicate applications
    if Application.objects.filter(job=job, applicant=seeker).exists():
        messages.info(request, "You already applied for this job.")
        return redirect("jobseeker_dashboard")

    # create application
    Application.objects.create(job=job, applicant=seeker)

    # notify employer
    Notification.objects.create(
        receiver=job.employer.user,
        message=f"{request.user.first_name} {request.user.last_name} applied for your job '{job.title}'."
    )

    # send email
    try:
        send_mail(
            subject=f"New Applicant for {job.title}",
            message=(
                f"Hi {job.employer.user.first_name},\n\n"
                f"You have a new applicant for '{job.title}'.\n"
                f"Applicant: {request.user.first_name} {request.user.last_name}\n\n"
                "Log in to BlueHire to review their profile."
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[job.employer.user.email],
            fail_silently=True,
        )
    except Exception:
        pass

    # messages.success(request, "You successfully applied for this job! PLEASE CHECK YOUR EMAIL OR SMS FOR FURTHER DETAILS")
    return redirect("jobseeker_dashboard")


# ====================================================
# 🔹 SYSTEM ADMIN DASHBOARD + VERIFY DOCS
# ====================================================
# Add this to your accounts/views.py - REPLACE the existing system_admin_dashboard function

@login_required
def system_admin_dashboard(request):
    if not hasattr(request.user, "role") or request.user.role != "Admin":
        messages.error(request, "Access denied.")
        return redirect("signin_register")

    jobseekers = JobSeeker.objects.select_related("user").all().order_by("-seeker_id")
    employers = Employer.objects.select_related("user").all().order_by("-employer_id")
    company_employers = employers.filter(employer_type="Company")
    personal_employers = employers.filter(employer_type="Personal")
    
    # ✅ Show last 30 days of logs by default (not all logs)
    thirty_days_ago = timezone.now() - timezone.timedelta(days=30)
    logs = SystemLog.objects.filter(
        action__isnull=False,
        timestamp__gte=thirty_days_ago
    ).order_by("-timestamp")[:50]  # limit to 50 most recent

    total_seekers = jobseekers.count()
    total_employers = employers.count()

    verified_seekers = jobseekers.filter(
        models.Q(approval_status="Approved") | models.Q(verification_status="Approved")
    ).count()
    verified_employers = employers.filter(
        models.Q(approval_status="Approved") | models.Q(is_verified=True)
    ).count()

    rejected_seekers = jobseekers.filter(approval_status="Rejected").count()
    rejected_employers = employers.filter(approval_status="Rejected").count()

    total_verified = verified_seekers + verified_employers
    total_rejected = rejected_seekers + rejected_employers

    pending_count = Job.objects.filter(is_approved=False).count()

    seeker_percent = round((verified_seekers / total_seekers * 100), 1) if total_seekers else 0
    employer_percent = round((verified_employers / total_employers * 100), 1) if total_employers else 0
    overall_percent = (
        round(((verified_seekers + verified_employers) /
               (total_seekers + total_employers) * 100), 1)
        if (total_seekers + total_employers) > 0 else 0
    )

    # ✅ FIXED: "Hired This Month" - tracks current calendar month only
    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Get applications from this month
    total_applications = Application.objects.filter(applied_date__gte=month_start).count()
    hired_applications = Application.objects.filter(
        applied_date__gte=month_start, 
        status="Hired"
    ).count()
    hired_percent = round((hired_applications / total_applications) * 100, 1) if total_applications else 0

    # ✅ NEW STATS: Jobs and Applications This Month
    jobs_this_month = Job.objects.filter(
        created_at__gte=month_start,
        is_approved=True
    ).count()
    
    applications_this_month = total_applications

    # ✅ BUILD AI ALERTS FROM ACTUAL JOBSEEKER DOCUMENTS
    alerts_data = []
    
    # Get all pending/suspicious jobseekers
    pending_seekers = jobseekers.filter(approval_status="Pending")
    
    for seeker in pending_seekers:
        user = seeker.user
        full_name = f"{user.first_name} {user.last_name}"
        
        # Check each document individually
        docs_to_check = [
            ("Barangay Clearance", seeker.barangay_clearance),
            ("Police Clearance", seeker.police_clearance),
            ("Nbi Clearance", seeker.nbi_clearance),
            ("Valid Id", seeker.valid_id),
        ]
        
        for doc_name, doc_file in docs_to_check:
            if doc_file:  # Only show if document exists
                # Get AI verification record
                verification = DocumentVerification.objects.filter(
                    user=user,
                    doc_type=doc_name
                ).first()
                
                confidence = 0.0
                status = "Pending"
                
                if verification:
                    confidence = float(verification.ai_confidence or 0.0)
                    status = verification.status
                
                alerts_data.append({
                    "user_type": "JobSeeker",
                    "username": user.username,
                    "full_name": full_name,
                    "doc_type": doc_name,
                    "file": doc_file,
                    "confidence": confidence,
                    "status": status,
                    "created_at": seeker.date_created,
                })
    
    # Get pending employers
    pending_employers = employers.filter(approval_status="Pending")
    
    for employer in pending_employers:
        user = employer.user
        full_name = f"{user.first_name} {user.last_name}"
        
        if employer.employer_type == "Company":
            docs_to_check = [
                ("Business Permit", employer.business_permit),
                ("Company ID", employer.company_id_file),
            ]
        else:
            docs_to_check = [
                ("Barangay Clearance", employer.barangay_clearance),
                ("Police Clearance", employer.police_clearance),
                ("Nbi Clearance", employer.nbi_clearance),
                ("Valid Id", employer.valid_id),
            ]
        
        for doc_name, doc_file in docs_to_check:
            if doc_file:
                verification = DocumentVerification.objects.filter(
                    user=user,
                    doc_type=doc_name
                ).first()
                
                confidence = 0.0
                status = "Pending"
                
                if verification:
                    confidence = float(verification.ai_confidence or 0.0)
                    status = verification.status
                
                alerts_data.append({
                    "user_type": "Employer",
                    "username": user.username,
                    "full_name": full_name,
                    "doc_type": doc_name,
                    "file": doc_file,
                    "confidence": confidence,
                    "status": status,
                    "created_at": employer.date_created if hasattr(employer, 'date_created') else timezone.now(),
                })

    context = {
        "jobseekers": jobseekers,
        "company_employers": company_employers,
        "personal_employers": personal_employers,
        "logs": logs,
        "seeker_percent": seeker_percent,
        "employer_percent": employer_percent,
        "overall_percent": overall_percent,
        "total_seekers": total_seekers,
        "total_employers": total_employers,
        "verified_seekers": verified_seekers,
        "verified_employers": verified_employers,
        "total_verified": total_verified,
        "total_rejected": total_rejected,
        "hired_percent": hired_percent,
        "hired_applications": hired_applications,  # ✅ NEW: actual number
        "total_applications": total_applications,  # ✅ NEW: total this month
        "jobs_this_month": jobs_this_month,  # ✅ NEW
        "applications_this_month": applications_this_month,  # ✅ NEW
        "pending_count": pending_count,
        "alerts_data": alerts_data,
        "BLUEHIRE_VERSION": "1.0.0",
    }
    return render(request, "core/system_admin_dashboard.html", context)


# ✅ NEW: Manual log clearing function
@login_required
def clear_old_logs(request):
    """Admin can manually clear logs with flexible time periods"""
    if not hasattr(request.user, "role") or request.user.role != "Admin":
        messages.error(request, "Access denied.")
        return redirect("signin_register")
    
    if request.method == "POST":
        days = int(request.POST.get("days", 90))
        
        if days == 0:
            # Clear ALL logs
            deleted_count = SystemLog.objects.all().delete()[0]
            message = f"Cleared all {deleted_count} log entries"
        else:
            # Clear logs older than specified days
            cutoff_date = timezone.now() - timezone.timedelta(days=days)
            deleted_count = SystemLog.objects.filter(timestamp__lt=cutoff_date).delete()[0]
            message = f"Cleared {deleted_count} logs older than {days} days"
        
        # Log this action
        SystemLog.objects.create(
            user=request.user,
            action=message,
            timestamp=timezone.now()
        )
        
        messages.success(request, f"✅ {message}.")
        return redirect("system_admin_dashboard")
    
    return redirect("system_admin_dashboard")


# ✅ FIXED: generate_report should NOT reset anything
@login_required
def generate_report(request):
    if not hasattr(request.user, "role") or request.user.role != "Admin":
        messages.error(request, "Access denied.")
        return redirect("signin_register")

    report_type = request.GET.get("type", "monthly_hires")

    wb = openpyxl.Workbook()
    ws = wb.active

    if report_type == "monthly_hires":
        ws.title = "Monthly Hires"
        ws.append(["Job Seeker", "Job Title", "Employer", "Hired Date", "Status"])
        
        # ✅ Get current month data (don't reset anything)
        now = timezone.now()
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        hires = Application.objects.filter(
            status="Hired", 
            applied_date__gte=month_start
        ).select_related('applicant__user', 'job__employer__user')
        
        for app in hires:
            ws.append([
                f"{app.applicant.user.first_name} {app.applicant.user.last_name}",
                str(app.job.title if app.job else "—"),
                str(app.job.employer.company_name if app.job and app.job.employer else "—"),
                app.applied_date.strftime("%Y-%m-%d"),
                app.status
            ])

    elif report_type == "verified_users":
        ws.title = "Verified Users"
        ws.append(["User Type", "Full Name", "Email", "Status", "Verified Date"])
        
        for s in JobSeeker.objects.filter(approval_status="Approved"):
            ws.append([
                "Job Seeker", 
                f"{s.user.first_name} {s.user.last_name}", 
                s.user.email, 
                "Approved",
                s.date_created.strftime("%Y-%m-%d") if s.date_created else "—"
            ])
        
        for e in Employer.objects.filter(is_verified=True):
            ws.append([
                "Employer", 
                f"{e.user.first_name} {e.user.last_name}", 
                e.user.email, 
                "Verified",
                "—"  # Add date_created to Employer model if needed
            ])

    elif report_type == "applications_overview":
        ws.title = "Applications Overview"
        ws.append(["Job Title", "Employer", "Total Applicants", "Hired", "Pending", "Rejected"])
        
        for job in Job.objects.all().select_related('employer__user'):
            total_apps = job.application_set.count()
            hired = job.application_set.filter(status="Hired").count()
            pending = job.application_set.filter(status="Pending").count()
            rejected = job.application_set.filter(status="Rejected").count()
            
            ws.append([
                job.title,
                job.employer.company_name or f"{job.employer.user.first_name} {job.employer.user.last_name}",
                total_apps,
                hired,
                pending,
                rejected
            ])

    elif report_type == "monthly_statistics":
        ws.title = "Monthly Statistics"
        now = timezone.now()
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        ws.append(["Metric", "Count"])
        ws.append(["Total Job Seekers", JobSeeker.objects.count()])
        ws.append(["Total Employers", Employer.objects.count()])
        ws.append(["Verified Job Seekers", JobSeeker.objects.filter(approval_status="Approved").count()])
        ws.append(["Verified Employers", Employer.objects.filter(is_verified=True).count()])
        ws.append(["Jobs Posted This Month", Job.objects.filter(created_at__gte=month_start).count()])
        ws.append(["Applications This Month", Application.objects.filter(applied_date__gte=month_start).count()])
        ws.append(["Hires This Month", Application.objects.filter(applied_date__gte=month_start, status="Hired").count()])

    # ✅ Log report generation (doesn't reset anything)
    SystemLog.objects.create(
        user=request.user,
        action=f"Generated {report_type} report",
        timestamp=timezone.now()
    )

    # Set headers and return file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"{report_type}_{datetime.date.today()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
@require_POST
def verify_document(request, user_type, user_id, doc_type):
    try:
        if user_type.lower() == "jobseeker":
            seeker = JobSeeker.objects.filter(user_id=user_id).first()
            if not seeker:
                messages.error(request, "Job seeker not found.")
                return redirect("system_admin_dashboard")

            decision = request.POST.get("action")
            if decision == "approve":
                seeker.verification_status = "Approved"
                seeker.user.is_verified = True
                DocumentVerification.objects.filter(user=seeker.user).update(status="Verified")
                msg = f"✅ Approved job seeker {seeker.user.username}'s documents"
            else:
                seeker.verification_status = "Rejected"
                seeker.user.is_verified = False
                DocumentVerification.objects.filter(user=seeker.user).update(status="Rejected")
                msg = f"❌ Rejected job seeker {seeker.user.username}'s documents"

            seeker.save()
            seeker.user.save()

        elif user_type.lower() == "employer":
            employer = Employer.objects.filter(user_id=user_id).first()
            if not employer:
                messages.error(request, "Employer not found.")
                return redirect("system_admin_dashboard")

            decision = request.POST.get("action")
            if decision == "approve":
                employer.is_verified = True
                employer.user.is_verified = True

                # ✅ mark all employer docs verified
                if employer.employer_type == "Company":
                    employer.business_permit_status = "Verified"
                    employer.company_id_status = "Verified"
                else:
                    employer.barangay_status = "Verified"
                    employer.police_status = "Verified"
                    employer.nbi_status = "Verified"
                    employer.valid_id_status = "Verified"

                msg = f"✅ Approved employer {employer.user.username}'s documents"
            else:
                employer.is_verified = False
                employer.user.is_verified = False

                if employer.employer_type == "Company":
                    employer.business_permit_status = "Rejected"
                    employer.company_id_status = "Rejected"
                else:
                    employer.barangay_status = "Rejected"
                    employer.police_status = "Rejected"
                    employer.nbi_status = "Rejected"
                    employer.valid_id_status = "Rejected"

                msg = f"❌ Rejected employer {employer.user.username}'s documents"

            employer.save()
            employer.user.save()

        else:
            messages.error(request, "Invalid user type.")
            return redirect("system_admin_dashboard")

        # ✅ create a system log
        SystemLog.objects.create(user=request.user, action=msg, timestamp=timezone.now())

        # ✅ create a notification
        Notification.objects.create(receiver=employer.user if user_type == "employer" else seeker.user,
                                    message=msg)

        messages.success(request, msg)

    except Exception as e:
        messages.error(request, f"Error verifying document: {e}")

    return redirect("system_admin_dashboard")


@login_required
def toggle_verification(request, user_id):
    try:
        user = Users.objects.get(pk=user_id)
        user.is_verified = not user.is_verified
        user.save()
        SystemLog.objects.create(
            user=request.user,
            action=f"Toggled verification for {user.username}",
            timestamp=timezone.now()
        )
        messages.success(request, f"{user.username}'s verification status updated.")
    except Users.DoesNotExist:
        messages.error(request, "User not found.")
    return redirect("system_admin_dashboard")

@login_required
def pending_jobs(request):
    """List of jobs waiting for admin approval"""
    if request.user.role != "Admin":
        messages.error(request, "Access denied.")
        return redirect("signin_register")

    # ✅ FIXED: Load employer and user data with the jobs
    pending_jobs = Job.objects.filter(
        is_approved=False
    ).select_related(
        'employer',           # Load the Employer relation
        'employer__user'      # Load the User relation through Employer
    ).order_by('-created_at')
    
    return render(request, "core/pending_jobs.html", {"pending_jobs": pending_jobs})

@login_required
@require_POST
def approve_job(request, job_id):
    """Approve a pending job posting"""
    from django.utils import timezone
    from jobs.models import Notification
    from systemlogs.models import SystemLog
    from jobs.models import Job

    # ✅ only admins can approve
    if request.user.role != "Admin":
        messages.error(request, "Access denied.")
        return redirect("signin_register")

    job = get_object_or_404(Job, pk=job_id)
    job.is_approved = True
    job.save()

    # ✅ log the action
    SystemLog.objects.create(
        user=request.user,
        action=f"Approved job '{job.title}' by {job.employer.company_name or job.employer.user.username}",
        timestamp=timezone.now(),
    )

    # ✅ notify employer
    Notification.objects.create(
        receiver=job.employer.user,
        message=f"✅ Your job posting '{job.title}' has been approved and is now live on BlueHire.",
        created_at=timezone.now(),
        is_read=False
    )

    messages.success(request, f"✅ Job '{job.title}' approved successfully and employer notified.")
    return redirect("pending_jobs")

@login_required
@require_POST  # ✅ ADD THIS
def reject_job(request, job_id):
    if not hasattr(request.user, "role") or request.user.role != "Admin":
        messages.error(request, "Access denied.")
        return redirect("homepage")

    job = get_object_or_404(Job, id=job_id)
    employer = job.employer
    job_title = job.title  # ✅ Save title before deleting

    # ✅ Create notification for employer
    Notification.objects.create(
        receiver=employer.user,
        message=f"❌ Your job posting '{job_title}' was rejected by the admin. Please review our posting guidelines and try again.",
        created_at=timezone.now(),
        is_read=False
    )

    # ✅ Log the action
    SystemLog.objects.create(
        user=request.user,
        action=f"Rejected and deleted job '{job_title}' by {employer.user.username}",
        timestamp=timezone.now()
    )

    # ✅ Delete the job
    job.delete()
    
    messages.success(request, f"Job '{job_title}' rejected, deleted, and employer notified.")
    return redirect("pending_jobs")

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.shortcuts import get_object_or_404
from accounts.models import JobSeeker, SystemLog
from jobs.models import Notification
from systemlogs.models import AIDocumentCheck
from accounts.models import DocumentVerification


@login_required
def approve_jobseeker(request, seeker_id):
    seeker = get_object_or_404(JobSeeker, seeker_id=seeker_id)
    seeker.approval_status = "Approved"
    seeker.verification_status = "Approved"
    seeker.save(update_fields=["approval_status", "verification_status"])

    # ✅ Mark ALL docs verified - INCLUDING BARANGAY CLEARANCE
    DocumentVerification.objects.filter(user=seeker.user).update(status="Verified")
    
    # ✅ Also create verification records if they don't exist for uploaded documents
    doc_types = [
        ("Barangay Clearance", seeker.barangay_clearance),
        ("Police Clearance", seeker.police_clearance),
        ("Nbi Clearance", seeker.nbi_clearance),
        ("Valid Id", seeker.valid_id),
    ]
    
    for doc_name, doc_file in doc_types:
        if doc_file:  # Only create if document exists
            DocumentVerification.objects.update_or_create(
                user=seeker.user,
                doc_type=doc_name,
                defaults={
                    "file": doc_file,
                    "status": "Verified",
                    "ai_confidence": 1.0,
                }
            )

    Notification.objects.create(
        receiver=seeker.user,
        message="✅ Your documents have been verified by the admin!"
    )
    SystemLog.objects.create(
        user=request.user,
        action=f"Approved jobseeker {seeker.user.username}",
        timestamp=timezone.now()
    )
    return JsonResponse({"ok": True})


@login_required
def reject_jobseeker(request, seeker_id):
    seeker = get_object_or_404(JobSeeker, seeker_id=seeker_id)
    seeker.approval_status = "Rejected"
    seeker.verification_status = "Rejected"
    seeker.save(update_fields=["approval_status", "verification_status"])

    # ✅ Mark all docs rejected - INCLUDING BARANGAY CLEARANCE
    DocumentVerification.objects.filter(user=seeker.user).update(status="Rejected")
    
    # ✅ Create verification records if they don't exist
    doc_types = [
        ("Barangay Clearance", seeker.barangay_clearance),
        ("Police Clearance", seeker.police_clearance),
        ("Nbi Clearance", seeker.nbi_clearance),
        ("Valid Id", seeker.valid_id),
    ]
    
    for doc_name, doc_file in doc_types:
        if doc_file:
            DocumentVerification.objects.update_or_create(
                user=seeker.user,
                doc_type=doc_name,
                defaults={
                    "file": doc_file,
                    "status": "Rejected",
                    "ai_confidence": 0.0,
                }
            )

    Notification.objects.create(
        receiver=seeker.user,
        message="❌ Your verification has been rejected. Please reupload valid documents."
    )
    SystemLog.objects.create(
        user=request.user,
        action=f"Rejected jobseeker {seeker.user.username}",
        timestamp=timezone.now()
    )
    return JsonResponse({"ok": True})


@login_required
def reupload_jobseeker(request, seeker_id):
    seeker = get_object_or_404(JobSeeker, seeker_id=seeker_id)
    seeker.approval_status = "Re-upload Required"
    seeker.verification_status = "Re-upload Required"
    seeker.save(update_fields=["approval_status", "verification_status"])

    # ✅ Reset all docs to pending - INCLUDING BARANGAY CLEARANCE
    DocumentVerification.objects.filter(user=seeker.user).update(status="Pending")
    
    # ✅ Create verification records if they don't exist
    doc_types = [
        ("Barangay Clearance", seeker.barangay_clearance),
        ("Police Clearance", seeker.police_clearance),
        ("Nbi Clearance", seeker.nbi_clearance),
        ("Valid Id", seeker.valid_id),
    ]
    
    for doc_name, doc_file in doc_types:
        if doc_file:
            DocumentVerification.objects.update_or_create(
                user=seeker.user,
                doc_type=doc_name,
                defaults={
                    "file": doc_file,
                    "status": "Pending",
                    "ai_confidence": 0.5,
                }
            )

    Notification.objects.create(
        receiver=seeker.user,
        message="⚠️ Please reupload clearer copies of your documents."
    )
    SystemLog.objects.create(
        user=request.user,
        action=f"Requested reupload from jobseeker {seeker.user.username}",
        timestamp=timezone.now()
    )
    return JsonResponse({"ok": True})

@login_required
def view_document(request, seeker_id, doc_type):
    """System Admin can view uploaded jobseeker documents directly"""
    seeker = get_object_or_404(JobSeeker, id=seeker_id)

    doc_map = {
        "barangay": seeker.barangay_clearance,
        "police": seeker.police_clearance,
        "nbi": seeker.nbi_clearance,
        "id": seeker.valid_id,
    }

    file_field = doc_map.get(doc_type.lower())
    if not file_field:
        messages.error(request, "Document not found or invalid type.")
        return redirect("system_admin_dashboard")

    # ensure file exists
    file_url = file_field.url if file_field else None
    if not file_url:
        messages.error(request, f"{doc_type.title()} not uploaded.")
        return redirect("system_admin_dashboard")

    return render(
        request,
        "core/view_document.html",
        {"file_url": file_url, "doc_type": doc_type.title(), "seeker": seeker},
    )


# ====================================================
# 🔹 EMPLOYER DASHBOARD + PROFILE
# ====================================================
from jobs.ai_matcher import rank_applicants
from accounts.models import Employer
from jobs.models import Notification

@login_required
def employer_dashboard(request):
    try:
        employer = Employer.objects.get(user=request.user)
    except Employer.DoesNotExist:
        messages.error(request, "Employer profile not found.")
        return redirect("signin_register")

    from django.db.models import Count, Q, Avg
    from jobs.models import Job
    from applications.models import Application
    from ratings.models import Rating
    
    jobs = Job.objects.filter(employer=employer).order_by("-created_at")
    
    jobs_with_stats = []
    for job in jobs:
        job.total_applicants = Application.objects.filter(job=job).count()
        job.hired_count = Application.objects.filter(
            job=job, status__in=["Hired", "Completed", "Not Completed"]
        ).count()

        hired_apps = Application.objects.filter(
            job=job, status__in=["Hired", "Completed", "Not Completed"]
        ).select_related("applicant__user")

        for app in hired_apps:
            existing_rating = Rating.objects.filter(
                reviewer=request.user,
                application=app,
                rated_jobseeker=app.applicant,
            ).first()
            app.existing_rating = existing_rating

        job.hired_applications = hired_apps
        job.rejected_count = Application.objects.filter(job=job, status="Rejected").count()
        job.pending_count = Application.objects.filter(job=job, status="Pending").count()
        job.approved_count = Application.objects.filter(job=job, status="Approved").count()

        for app in job.application_set.all():
            if app.applicant:
                avg = Rating.objects.filter(
                    rated_jobseeker=app.applicant, is_visible=True
                ).aggregate(avg=Avg('score'))['avg']
                app.applicant.avg_rating = round(avg, 1) if avg else None

        jobs_with_stats.append(job)

    total_jobs = len(jobs_with_stats)
    total_applicants = Application.objects.filter(job__employer=employer).count()
    pending_verifications = 0 if employer.is_verified else 1

    fields = [
        employer.company_name,
        employer.company_description,
        employer.company_logo,
        employer.business_permit,
    ]
    completed = sum(1 for f in fields if f)
    profile_completion = int((completed / len(fields)) * 100) if fields else 0

    recent_applicants = (
        Application.objects.filter(
            job__employer=employer,
            status__in=["Hired", "Completed", "Not Completed"]
        )
        .select_related("job", "applicant__user")
        .order_by("-applied_date")[:5]
    )

    # Attach existing rating to recent applicants too
    for app in recent_applicants:
        existing_rating = Rating.objects.filter(
            reviewer=request.user,
            application=app,
            rated_jobseeker=app.applicant,
        ).first()
        app.existing_rating = existing_rating

    notifications = Notification.objects.filter(
        receiver=employer.user
    ).order_by("-created_at")[:10]
    
    unread_count = Notification.objects.filter(
        receiver=employer.user,
        is_read=False
    ).count()

    return render(
        request,
        "core/employer_dashboard.html",
        {
            "employer": employer,
            "jobs": jobs_with_stats,
            "total_jobs": total_jobs,
            "total_applicants": total_applicants,
            "pending_verifications": pending_verifications,
            "profile_completion": profile_completion,
            "notifications": notifications,
            "unread_count": unread_count,
            "recent_applicants": recent_applicants,
        },
    )


@login_required
def upload_employer_documents(request):
    employer = Employer.objects.filter(user=request.user).first()
    if not employer:
        messages.error(request, "Employer profile not found.")
        return redirect("employer_dashboard")

    if request.method == "POST":
        uploaded_files = {}

        # =============== COMPANY EMPLOYERS ===============
        if employer.employer_type == "Company":
            # Business Permit
            if "business_permit" in request.FILES:
                employer.business_permit = request.FILES["business_permit"]
                employer.business_permit_status = "Pending"
                uploaded_files["business_permit"] = request.FILES["business_permit"]

            # Company ID
            if "company_id_file" in request.FILES:
                employer.company_id_file = request.FILES["company_id_file"]
                employer.company_id_status = "Pending"
                uploaded_files["company_id_file"] = request.FILES["company_id_file"]

        # =============== PERSONAL EMPLOYERS ===============
        else:
            # Barangay Clearance
            if "barangay_clearance" in request.FILES:
                employer.barangay_clearance = request.FILES["barangay_clearance"]
                employer.barangay_status = "Pending"
                uploaded_files["barangay_clearance"] = request.FILES["barangay_clearance"]

            # Police Clearance
            if "police_clearance" in request.FILES:
                employer.police_clearance = request.FILES["police_clearance"]
                employer.police_status = "Pending"
                uploaded_files["police_clearance"] = request.FILES["police_clearance"]

            # NBI Clearance
            if "nbi_clearance" in request.FILES:
                employer.nbi_clearance = request.FILES["nbi_clearance"]
                employer.nbi_status = "Pending"
                uploaded_files["nbi_clearance"] = request.FILES["nbi_clearance"]

            # Valid ID
            if "valid_id" in request.FILES:
                employer.valid_id = request.FILES["valid_id"]
                employer.valid_id_status = "Pending"
                uploaded_files["valid_id"] = request.FILES["valid_id"]
            
            # Valid ID Back
            if "valid_id_back" in request.FILES:
                employer.valid_id_back = request.FILES["valid_id_back"]
                uploaded_files["valid_id_back"] = request.FILES["valid_id_back"]

            # ✅ Business Permit (OPTIONAL for Personal/SME - Auto-verified, not reviewed by admin)
            if "business_permit" in request.FILES:
                employer.business_permit = request.FILES["business_permit"]
                employer.business_permit_status = "Verified"  # ✅ Auto-verified, not reviewed
                uploaded_files["business_permit"] = request.FILES["business_permit"]

            # ID Type
            valid_id_type = request.POST.get("valid_id_type")
            if valid_id_type:
                employer.valid_id_type = valid_id_type

        # ✅ Save first
        employer.save()

        # ✅ AI verification
        run_ai_verification_for_user(request.user, uploaded_files)

        # ✅ NEW: CHECK IF ALL REQUIRED DOCUMENTS ARE UPLOADED
        if employer.employer_type == "Company":
            all_docs_uploaded = all([
                employer.business_permit,
                employer.company_id_file
            ])
        else:
            all_docs_uploaded = all([
                employer.barangay_clearance,
                employer.police_clearance,
                employer.nbi_clearance,
                employer.valid_id
            ])

        # ✅ ONLY SET TO PENDING IF ALL DOCS ARE UPLOADED
        if all_docs_uploaded:
            employer.approval_status = "Pending"
            
            # ✅ Notify admins
            from django.contrib.auth import get_user_model
            from jobs.models import Notification
            from django.utils import timezone

            User = get_user_model()
            admins = User.objects.filter(role="Admin")
            for admin in admins:
                Notification.objects.create(
                    receiver=admin,
                    message=f"{request.user.first_name} {request.user.last_name} (Employer) has uploaded all documents and is ready for verification.",
                    created_at=timezone.now(),
                )
            
            messages.success(request, "✅ All documents uploaded! Your application is now pending admin review.")
        else:
            employer.approval_status = "Incomplete"
            messages.success(request, "Documents uploaded. Please upload all required documents to proceed.")
        
        employer.save()

        update_employer_progress(employer)
        return redirect("employer_profile")

    return redirect("employer_profile")


def update_employer_progress(employer):
    """Recalculate employer profile completion with better coverage."""
    if employer.employer_type == "Company":
        fields = [
            employer.company_name,
            employer.company_description,
            employer.company_address,
            employer.company_logo,
            employer.business_permit,
            employer.company_id_file,
        ]
    else:
        fields = [
            employer.personal_address,
            employer.bio,
            employer.profile_photo,
            employer.barangay_clearance,
            employer.police_clearance,
            employer.nbi_clearance,
            employer.valid_id,  # ✅ added to include ID in completion
        ]

    filled = sum(1 for f in fields if f)
    employer.profile_completion = int((filled / len(fields)) * 100)
    employer.save(update_fields=["profile_completion"])

    
    
@login_required
def employer_profile(request):
    employer = Employer.objects.filter(user=request.user).first()  # ✅ FIXED

    # ✅ if employer not found, don't log out — redirect to dashboard instead
    if not employer:
        messages.warning(request, "Employer profile temporarily unavailable. Please reload.")
        return redirect("employer_dashboard")

    # ✅ ensure employer_type always has a default
    if not employer.employer_type:
        employer.employer_type = "Personal"
        employer.save()

    # ✅ safely refresh completion bar
    try:
        update_employer_progress(employer)
    except Exception as e:
        print("Profile completion error:", e)

    employer.refresh_from_db()

    return render(
        request,
        "core/employer_profile.html",
        {"employer": employer, "completion": employer.profile_completion},
    )
    
    
@login_required
def update_employer_profile(request):
    employer = Employer.objects.filter(user=request.user).first()
    if not employer:
        messages.error(request, "Employer profile not found.")
        return redirect("employer_dashboard")

    user = request.user  # for updating name and phone

    if request.method == "POST":
        # ✅ UPDATE USER MODEL FIELDS
        user.first_name = request.POST.get("first_name", "").strip()
        user.last_name = request.POST.get("last_name", "").strip()
        user.phone = request.POST.get("phone", "").strip()
        user.save()

        # ✅ company employer fields
        if employer.employer_type == "Company":
            employer.company_name = request.POST.get("company_name", employer.company_name)
            employer.company_description = request.POST.get("company_description", employer.company_description)
            employer.company_address = request.POST.get("company_address", employer.company_address)

            if "company_logo" in request.FILES:
                employer.company_logo = request.FILES["company_logo"]
            if "business_permit" in request.FILES:
                employer.business_permit = request.FILES["business_permit"]
            if "company_id_file" in request.FILES:
                employer.company_id_file = request.FILES["company_id_file"]

        # ✅ personal employer fields
        else:
            employer.personal_address = request.POST.get("personal_address", employer.personal_address)
            employer.bio = request.POST.get("bio", employer.bio)

            if "profile_photo" in request.FILES:
                employer.profile_photo = request.FILES["profile_photo"]
            # ... rest of document uploads

        employer.save()
        update_employer_progress(employer)

        messages.success(request, "Profile updated successfully!")
        return redirect("employer_profile")

    employer.refresh_from_db()
    return render(request, "core/employer_profile.html", {"employer": employer})



def employer_public_profile(request, employer_id):
    from ratings.models import Rating
    from django.db.models import Count

    employer = get_object_or_404(Employer, employer_id=employer_id)
    posted_jobs = Job.objects.filter(employer=employer)

    if employer.employer_type == "Company":
        documents = {
            "Business Permit": employer.business_permit,
            "Company ID": employer.company_id_file,
            "Barangay Clearance": employer.barangay_clearance,
            "Police Clearance": employer.police_clearance,
            "NBI Clearance": employer.nbi_clearance,
        }
    else:
        documents = {
            "Valid ID": employer.company_id_file,
            "Barangay Clearance": employer.barangay_clearance,
            "Police Clearance": employer.police_clearance,
            "NBI Clearance": employer.nbi_clearance,
        }

    ratings = Rating.objects.filter(
        rated_employer=employer, is_visible=True
    ).select_related('reviewer', 'application__job')
    stats = ratings.aggregate(avg=Avg('score'), total=Count('id'))
    avg_score = round(stats['avg'] or 0, 1)
    avg_rating = avg_score if stats['total'] > 0 else None
    total_ratings = stats['total']
    distribution = {i: ratings.filter(score=i).count() for i in range(5, 0, -1)}

    return render(
        request,
        "core/employer_public_profile.html",
        {
            "employer": employer,
            "documents": documents,
            "jobs": posted_jobs,
            "ratings": ratings,
            "avg_rating": avg_rating,
            "total_ratings": total_ratings,
            "distribution": distribution,
            "filled_stars": range(1, round(avg_score) + 1),
            "empty_stars": range(round(avg_score) + 1, 6),
        },
    )


@login_required
def view_jobseeker_profile(request, seeker_id):
    seeker = get_object_or_404(JobSeeker, id=seeker_id)
    context = {
        "seeker": seeker,
        "user": seeker.user  # so template sees {{ user }} properly
    }
    return render(request, "core/jobseeker_profile.html", context)


@login_required
def employer_public_profile(request, employer_id):
    # get employer record
    employer = get_object_or_404(Employer, employer_id=employer_id)

    # posted jobs by employer
    posted_jobs = Job.objects.filter(employer=employer)

    # decide which docs to show
    if employer.employer_type == "Company":
        documents = {
            "Business Permit": employer.business_permit,
            "Company ID": employer.company_id_file,
            "Barangay Clearance": employer.barangay_clearance,
            "Police Clearance": employer.police_clearance,
            "NBI Clearance": employer.nbi_clearance,
        }
    else:
        documents = {
            "Valid ID": employer.company_id_file,
            "Barangay Clearance": employer.barangay_clearance,
            "Police Clearance": employer.police_clearance,
            "NBI Clearance": employer.nbi_clearance,
        }

    return render(
        request,
        "core/employer_public_profile.html",
        {
            "employer": employer,
            "documents": documents,
            "jobs": posted_jobs,
        },
    )
 
    
from django.shortcuts import render, get_object_or_404
from django.db.models import Q
from accounts.models import Employer, JobSeeker, JobSeekerCertificate
from jobs.models import Job
from applications.models import Application

@login_required
def view_applications(request):
    """
    Employer view to see all applications with AI-ranked applicants
    """
    
    try:
        employer = Employer.objects.get(user=request.user)
    except Employer.DoesNotExist:
        messages.error(request, "Employer profile not found.")
        return redirect("signin_register")

    from django.db.models import Count, Q, Avg
    from jobs.models import Job
    from applications.models import Application
    from accounts.models import JobSeekerCertificate
    from ratings.models import Rating
    
    # Get all jobs for this employer
    jobs = Job.objects.filter(employer=employer).prefetch_related(
        "job_applications__applicant__user"
    ).order_by("-created_at")

    # Calculate stats for each job
    jobs_with_stats = []
    for job in jobs:
        job.total_applicants = Application.objects.filter(job=job).count()
        job.hired_count = Application.objects.filter(job=job, status="Hired").count()
        job.rejected_count = Application.objects.filter(job=job, status="Rejected").count()
        job.pending_count = Application.objects.filter(job=job, status="Pending").count()
        job.approved_count = Application.objects.filter(job=job, status="Approved").count()
        jobs_with_stats.append(job)

    total_jobs = len(jobs_with_stats)
    total_applicants = Application.objects.filter(job__employer=employer).count()
    pending_verifications = 0 if employer.is_verified else 1

    # Profile completion
    fields = [
        employer.company_name,
        employer.company_description,
        employer.company_logo,
        employer.business_permit,
    ]
    completed = sum(1 for f in fields if f)
    profile_completion = int((completed / len(fields)) * 100) if fields else 0

    # Recent hires
    recent_applicants = (
        Application.objects.filter(
            job__employer=employer,
            status__in=["Hired", "Completed", "Not Completed"]
        )
        .select_related("job", "applicant__user")
        .order_by("-applied_date")[:5]
    )

    # Notifications
    notifications = Notification.objects.filter(
        receiver=employer.user
    ).order_by("-created_at")[:10]
    
    unread_count = Notification.objects.filter(
        receiver=employer.user,
        is_read=False
    ).count()

    # ✅ BUILD APPLICANT RATINGS MAP
    applicant_ratings = {}
    all_seeker_ids = []
    for job in jobs_with_stats:
        for app in job.job_applications.all():
            if app.applicant:
                all_seeker_ids.append(app.applicant.seeker_id)
    for seeker_id in set(all_seeker_ids):
        avg = Rating.objects.filter(
            rated_jobseeker__seeker_id=seeker_id, is_visible=True
        ).aggregate(avg=Avg('score'))['avg']
        if avg:
            applicant_ratings[seeker_id] = round(avg, 1)

    # ✅ AI RECOMMENDATIONS
    ai_recommendations = {}
    
    print("\n🎯 RANKING APPLICANTS (NO SKILLS, ONLY CERTS + PREFERENCES + EXPERIENCE + ABOUT)")
    print("="*80)
    
    for job in jobs_with_stats:
        print(f"\n📌 Job: {job.title} | Category: {job.category}")
        
        applications = Application.objects.filter(
            job=job
        ).exclude(
            status="Hired"
        ).select_related('applicant', 'applicant__user')
        
        applicants = [app.applicant for app in applications if app.applicant]
        
        print(f"   👥 {len(applicants)} applicants (excluding hired)")
        
        if not applicants:
            print(f"   ⚠️ No applicants to rank")
            ai_recommendations[job.id] = []
            continue
        
        rankings = []
        
        for seeker in applicants:
            score = 0
            match_reason = []
            
            uploaded_certs = JobSeekerCertificate.objects.filter(jobseeker=seeker)
            has_uploaded_certificates = uploaded_certs.exists()
            
            if has_uploaded_certificates:
                score += 30
                match_reason.append(f"Has {uploaded_certs.count()} NC2/certificates (+30)")
                print(f"      ✅ {seeker.user.username}: +30 pts ({uploaded_certs.count()} certificates)")
            
            if hasattr(seeker, 'certifications') and seeker.certifications:
                certs_list = seeker.certifications if isinstance(seeker.certifications, list) else []
                if certs_list:
                    score += 10
                    match_reason.append("Additional certifications (+10)")
                    print(f"      ✅ {seeker.user.username}: +10 pts (certifications JSONField)")
            
            if seeker.preferred_job_types:
                job_types_list = seeker.preferred_job_types if isinstance(seeker.preferred_job_types, list) else [str(seeker.preferred_job_types)]
                
                if job.category in job_types_list:
                    score += 40
                    match_reason.append("Exact category in preferences (+40)")
                    print(f"      ✅ {seeker.user.username}: +40 pts (exact category match)")
                else:
                    job_category_lower = job.category.lower()
                    for pref in job_types_list:
                        pref_lower = str(pref).lower()
                        if pref_lower in job_category_lower or job_category_lower in pref_lower:
                            score += 25
                            match_reason.append("Similar category match (+25)")
                            print(f"      ✅ {seeker.user.username}: +25 pts (similar category)")
                            break
            
            if hasattr(seeker, 'experience_years') and seeker.experience_years:
                try:
                    years = int(seeker.experience_years)
                    experience_score = min(years / 5.0, 1.0) * 20
                    score += experience_score
                    match_reason.append(f"{years} years experience (+{experience_score:.1f})")
                    print(f"      ✅ {seeker.user.username}: +{experience_score:.1f} pts ({years} years)")
                except:
                    pass
            
            if hasattr(seeker, 'about') and seeker.about and len(seeker.about.strip()) > 20:
                score += 5
                match_reason.append("Complete profile with about (+5)")
                print(f"      ✅ {seeker.user.username}: +5 pts (has about section)")
            
            score = min(100, round(score, 1))
            
            if score >= 90:
                match_level = "Excellent Match"
            elif score >= 70:
                match_level = "Good Match"
            elif score >= 50:
                match_level = "Fair Match"
            else:
                match_level = "Weak Match"
            
            print(f"      📊 {seeker.user.username}: {score}% - {match_level}")
            print(f"         Reasons: {', '.join(match_reason) if match_reason else 'No matching criteria'}")
            
            rankings.append({
                "seeker": seeker,
                "score": score,
                "match_level": match_level,
                "uploaded_certificates": uploaded_certs,
            })
        
        rankings.sort(key=lambda x: x['score'], reverse=True)
        
        print(f"   ✅ Ranked {len(rankings)} applicants")
        
        ai_recommendations[job.id] = rankings[:10]

    print("\n" + "="*80)
    print(f"🎯 FINAL RESULTS:")
    for job_id, recs in ai_recommendations.items():
        print(f"   Job {job_id}: {len(recs)} ranked applicants")
        for rec in recs[:3]:
            print(f"      - {rec['seeker'].user.username}: {rec['score']}% ({rec['match_level']})")
    print("="*80 + "\n")

    context = {
        "jobs": jobs_with_stats,
        "ai_recommendations": ai_recommendations,
        "employer": employer,
        "total_jobs": total_jobs,
        "total_applicants": total_applicants,
        "pending_verifications": pending_verifications,
        "profile_completion": profile_completion,
        "notifications": notifications,
        "unread_count": unread_count,
        "recent_applicants": recent_applicants,
        "applicant_ratings": applicant_ratings,   # ← NEW
    }
    
    return render(request, "core/employer_applications.html", context)


@login_required
def approve_employer(request, emp_id):
    print(f"🔴 APPROVE EMPLOYER CALLED - emp_id: {emp_id}")
    e = get_object_or_404(Employer, employer_id=emp_id)
    print(f"🔴 Found employer: {e.user.username}, Type: {e.employer_type}")
    print(f"🔴 BEFORE - business_permit_status: {e.business_permit_status}")
    
    e.approval_status = "Approved"
    e.is_verified = True
    
    if e.employer_type == "Company":
        e.business_permit_status = "Approved"
        e.company_id_status = "Approved"
        print(f"🔴 SET Company statuses to Approved")
    else:
        e.barangay_status = "Approved"
        e.police_status = "Approved"
        e.nbi_status = "Approved"
        e.valid_id_status = "Approved"
        print(f"🔴 SET Personal statuses to Approved")
    
    e.save()
    print(f"🔴 AFTER SAVE - business_permit_status: {e.business_permit_status}")
    
    # Rest of the function...
    DocumentVerification.objects.filter(user=e.user).update(status="Verified")
    Notification.objects.create(
        receiver=e.user,
        message="✅ Your employer documents were verified and approved!"
    )
    SystemLog.objects.create(
        user=request.user,
        action=f"Approved employer {e.user.username}",
        timestamp=timezone.now()
    )
    return JsonResponse({"ok": True})


@login_required
def reject_employer(request, emp_id):
    e = get_object_or_404(Employer, employer_id=emp_id)
    e.approval_status = "Rejected"
    e.is_verified = False
    
    # ✅ FIX: Set status to "Rejected"
    if e.employer_type == "Company":
        e.business_permit_status = "Rejected"
        e.company_id_status = "Rejected"
    else:
        e.barangay_status = "Rejected"
        e.police_status = "Rejected"
        e.nbi_status = "Rejected"
        e.valid_id_status = "Rejected"
    
    e.save()

    # ✅ Mark all docs rejected in DocumentVerification table
    DocumentVerification.objects.filter(user=e.user).update(status="Rejected")

    Notification.objects.create(
        receiver=e.user,
        message="❌ Your employer verification was rejected. Please reupload your files."
    )
    SystemLog.objects.create(
        user=request.user,
        action=f"Rejected employer {e.user.username}",
        timestamp=timezone.now()
    )
    return JsonResponse({"ok": True})


@login_required
def reupload_employer(request, emp_id):
    e = get_object_or_404(Employer, employer_id=emp_id)
    e.approval_status = "Re-upload Required"
    e.is_verified = False
    
    # ✅ FIX: Set status to "Pending" for reupload
    if e.employer_type == "Company":
        e.business_permit_status = "Pending"
        e.company_id_status = "Pending"
    else:
        e.barangay_status = "Pending"
        e.police_status = "Pending"
        e.nbi_status = "Pending"
        e.valid_id_status = "Pending"
    
    e.save()

    # ✅ Reset all docs to pending in DocumentVerification table
    DocumentVerification.objects.filter(user=e.user).update(status="Pending")

    Notification.objects.create(
        receiver=e.user,
        message="⚠️ Please reupload clearer copies of your verification documents."
    )
    SystemLog.objects.create(
        user=request.user,
        action=f"Requested reupload from employer {e.user.username}",
        timestamp=timezone.now()
    )
    return JsonResponse({"ok": True})

@login_required
@csrf_exempt
def toggle_job_availability(request, job_id):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    try:
        data = json.loads(request.body)
        available = data.get("available", True)
    except Exception:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    job = Job.objects.filter(id=job_id, employer__user=request.user).first()
    if not job:
        return JsonResponse({"error": "Job not found or unauthorized"}, status=404)

    job.is_available = available
    job.save(update_fields=["is_available"])
    return JsonResponse({"ok": True, "status": "Available" if available else "Unavailable"})

# ====================================================
# 🔹 NOTIFICATIONS + APPLICATION STATUS
# ====================================================
@login_required
def update_application_status(request, app_id):
    """Approve, Hire, or Reject applications with auto-close support"""
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid request method"}, status=405)

    status = request.POST.get("status")
    if not status:
        return JsonResponse({"success": False, "error": "Missing status"}, status=400)

    print(f"[DEBUG] ========================================")
    print(f"[DEBUG] Received app_id: {app_id} (type: {type(app_id)})")
    print(f"[DEBUG] Requested status: {status}")
    print(f"[DEBUG] User: {request.user.username}")

    try:
        # Try to get the application - handles UUID strings
        app = Application.objects.select_related(
            "applicant__user", "job", "job__employer"
        ).get(application_id=app_id)
        
        print(f"[DEBUG] ✅ Found application for job: {app.job.title}")
        print(f"[DEBUG] Current status: {app.status}")
        print(f"[DEBUG] Applicant: {app.applicant.user.username if app.applicant else 'None'}")
        
    except Application.DoesNotExist:
        print(f"[ERROR] ❌ Application not found with ID: {app_id}")
        return JsonResponse({"success": False, "error": f"Application not found"}, status=404)
    except Exception as e:
        print(f"[ERROR] ❌ Database error: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({"success": False, "error": f"Database error: {str(e)}"}, status=500)

    # Check if employer owns this job
    try:
        if not hasattr(app.job, 'employer'):
            print(f"[ERROR] ❌ Job has no employer")
            return JsonResponse({"success": False, "error": "Invalid job configuration"}, status=500)
        
        if not hasattr(app.job.employer, 'user'):
            print(f"[ERROR] ❌ Employer has no user")
            return JsonResponse({"success": False, "error": "Invalid employer configuration"}, status=500)
        
        if app.job.employer.user != request.user:
            print(f"[ERROR] ❌ Unauthorized: {request.user.username} tried to modify {app.job.employer.user.username}'s job")
            return JsonResponse({"success": False, "error": "Unauthorized"}, status=403)
        
        print(f"[DEBUG] ✅ Authorization passed")
        
    except Exception as e:
        print(f"[ERROR] ❌ Authorization check failed: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({"success": False, "error": f"Authorization error: {str(e)}"}, status=500)

    # Save old status to detect changes
    old_status = app.status
    
    # Update status
    try:
        app.status = status
        app.save()
        print(f"[SUCCESS] ✅ Updated app {app_id}: '{old_status}' → '{status}'")
    except Exception as e:
        print(f"[ERROR] ❌ Failed to save status: {e}")
        import traceback
        traceback.print_exc()
        return JsonResponse({"success": False, "error": f"Failed to update status: {str(e)}"}, status=500)

    # ✅ CHECK QUOTA AND AUTO-CLOSE (only when changing TO "Hired")
    job_closed = False
    hired_count = 0
    quota = 0
    
    if status == "Hired" and old_status != "Hired":
        try:
            job = app.job
            print(f"[DEBUG] Checking quota for job '{job.title}'")
            
            if job.auto_close_on_quota and job.is_available:
                hired_count = Application.objects.filter(job=job, status="Hired").count()
                quota = job.hiring_quota
                print(f"[DEBUG] Hired count: {hired_count}/{quota}")
                
                if hired_count >= quota:
                    job.is_available = False
                    job.save(update_fields=['is_available'])
                    job_closed = True
                    print(f"[SUCCESS] ✅ Job '{job.title}' auto-closed!")
                    
                    # Notify employer
                    Notification.objects.create(
                        receiver=request.user,
                        message=f"✅ Your job '{job.title}' has been automatically marked unavailable because you've hired {quota} worker(s).",
                    )
        except Exception as e:
            print(f"[WARNING] ⚠️ Auto-close check failed: {e}")
            # Don't fail the whole request if auto-close fails
            import traceback
            traceback.print_exc()

    # ✅ NOTIFY APPLICANT
    try:
        if status == "Approved":
            message = f"🎉 Congratulations! Your application for '{app.job.title}' has been approved!"
        elif status == "Hired":
            message = f"✅ You've been hired for '{app.job.title}'! Expect a message soon for onboarding."
            if job_closed:
                message += " This position is now filled."
        elif status == "Rejected":
            message = f"❌ Your application for '{app.job.title}' was not successful. Don't give up – more jobs await!"
        elif status == "Completed":
            message = f"🏁 Your job '{app.job.title}' has been marked as completed. You can now rate your employer!"
        elif status == "Not Completed":
            message = f"⚠️ Your application for '{app.job.title}' was marked as Not Completed by the employer."
        else:
            message = f"Your application status for '{app.job.title}' was updated to {status}."

        if app.applicant and app.applicant.user:
            Notification.objects.create(
                receiver=app.applicant.user,
                message=message,
            )
            print(f"[SUCCESS] ✅ Notification sent to {app.applicant.user.username}")
        else:
            print(f"[WARNING] ⚠️ No applicant user to notify")
            
    except Exception as e:
        print(f"[WARNING] ⚠️ Failed to send notification: {e}")
        # Don't fail the whole request if notification fails
        import traceback
        traceback.print_exc()

    # Calculate current stats
    try:
        current_hired = Application.objects.filter(job=app.job, status="Hired").count()
    except Exception as e:
        print(f"[WARNING] ⚠️ Failed to count hired: {e}")
        current_hired = 0

    print(f"[DEBUG] ========================================")

    # ✅ RETURN PROPER JSON RESPONSE
    return JsonResponse({
        "success": True, 
        "status": status,
        "job_closed": job_closed,
        "hired_count": current_hired,
        "quota": app.job.hiring_quota
    })

# ====================================================
# REPORTS
# ====================================================

@login_required
def generate_report(request):
    if not hasattr(request.user, "role") or request.user.role != "Admin":
        messages.error(request, "Access denied.")
        return redirect("signin_register")

    report_type = request.GET.get("type", "monthly_hires")

    wb = openpyxl.Workbook()
    ws = wb.active

    if report_type == "monthly_hires":
        ws.title = "Monthly Hires"
        ws.append(["Job Seeker", "Job Title", "Employer", "Hired Date"])
        month_start = timezone.now().replace(day=1)
        hires = Application.objects.filter(status="Hired", applied_date__gte=month_start)
        for app in hires:
            ws.append([
                str(app.applicant),
                str(app.job.title if app.job else "—"),
                str(app.job.employer.company_name if app.job and app.job.employer else "—"),
                app.applied_date.strftime("%Y-%m-%d")
            ])

    elif report_type == "verified_users":
        ws.title = "Verified Users"
        ws.append(["User Type", "Full Name", "Email", "Status"])
        for s in JobSeeker.objects.filter(verification_status="Approved"):
            ws.append(["Job Seeker", f"{s.user.first_name} {s.user.last_name}", s.user.email, "Approved"])
        for e in Employer.objects.filter(is_verified=True):
            ws.append(["Employer", f"{e.user.first_name} {e.user.last_name}", e.user.email, "Verified"])

    elif report_type == "applications_overview":
        ws.title = "Applications Overview"
        ws.append(["Job Title", "Total Applicants", "Hired"])
        from jobs.models import Job
        for job in Job.objects.all():
            total_apps = job.application_set.count()
            hired = job.application_set.filter(status="Hired").count()
            ws.append([job.title, total_apps, hired])

    # set headers
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"{report_type}_{datetime.date.today()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ====================================================
# 🔹 LOGOUT
# ====================================================
def logout_view(request):
    """Force logout with detailed debugging"""
    print("\n" + "="*60)
    print("🔴 LOGOUT ATTEMPT")
    print("="*60)
    print(f"User before logout: {request.user.username if request.user.is_authenticated else 'Anonymous'}")
    print(f"User role: {request.user.role if request.user.is_authenticated else 'None'}")
    print(f"Session key before: {request.session.session_key}")
    print(f"Is authenticated before: {request.user.is_authenticated}")
    
    try:
        # Store username before logout
        username = request.user.username if request.user.is_authenticated else "User"
        user_role = request.user.role if request.user.is_authenticated else None
        
        # Nuclear option - clear everything
        logout(request)
        print(f"✅ Django logout() called")
        
        request.session.flush()
        print(f"✅ Session flushed")
        
        # Force session to be deleted
        if hasattr(request, 'session'):
            try:
                request.session.delete()
                print(f"✅ Session deleted")
            except Exception as e:
                print(f"⚠️ Session delete error: {e}")
        
        print(f"User after logout: {request.user.username if request.user.is_authenticated else 'Anonymous'}")
        print(f"Is authenticated after: {request.user.is_authenticated}")
        print(f"Session key after: {request.session.session_key}")
        
        messages.success(request, f"{username} logged out successfully.")
        
    except Exception as e:
        print(f"❌ Logout error: {e}")
        import traceback
        traceback.print_exc()
    
    # Force redirect to login page
    print(f"🔄 Redirecting to signin_register...")
    print("="*60 + "\n")
    
    response = redirect("signin_register")
    
    # Clear all cookies related to authentication
    response.delete_cookie('sessionid')
    response.delete_cookie('csrftoken')
    
    # Prevent caching
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response

@login_required
def password_change(request):
    if request.method == "POST":
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # ✅ refresh session so user stays logged in
            update_session_auth_hash(request, user)
            messages.success(request, "Password changed successfully.")

            # redirect manually based on role
            if request.user.role == "Employer":
                return redirect("employer_profile")
            elif request.user.role == "JobSeeker":
                return redirect("jobseeker_profile")
            else:
                return redirect("homepage")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = PasswordChangeForm(request.user)

    return render(request, "core/password_change.html", {"form": form})


@login_required
def password_change_done(request):
    role = getattr(request.user, "role", None)
    if role == "Employer":
        return redirect("employer_dashboard")
    elif role == "JobSeeker":
        return redirect("jobseeker_dashboard")
    return redirect("homepage")

# ====================================================
#                     A I
# ====================================================
@login_required
def ai_recommend_seekers(request, job_id):
    job = get_object_or_404(Job, id=job_id, employer=request.user.employer)
    seekers = JobSeeker.objects.all()
    ranked = compute_similarity(seekers, job)

    # top 10
    top_seekers = ranked[:10]

    context = {
        "job": job,
        "ranked_seekers": top_seekers,
    }
    return render(request, "core/ai_recommendations.html", context)



from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
# simple reusable scoring function
def compute_match_score(seeker, job):
    """
    returns similarity percentage between jobseeker profile and job details
    """
    seeker_text = ""

    # combine skills and preferred job types
    if seeker.skills:
        seeker_text += " ".join(seeker.skills) if isinstance(seeker.skills, list) else str(seeker.skills)
    if hasattr(seeker, "preferred_job_types") and seeker.preferred_job_types:
        seeker_text += " " + (
            " ".join(seeker.preferred_job_types)
            if isinstance(seeker.preferred_job_types, list)
            else str(seeker.preferred_job_types)
        )

    job_text = f"{job.title} {job.category} {job.description}"

    if not seeker_text.strip() or not job_text.strip():
        return 0

    # TF-IDF and cosine similarity
    vectorizer = TfidfVectorizer(stop_words="english")
    vectors = vectorizer.fit_transform([seeker_text, job_text])
    similarity = cosine_similarity(vectors[0:1], vectors[1:2])[0][0]
    return round(similarity * 100, 1)


# Replace your jobseeker_recommendations view in accounts/views.py

@login_required
def jobseeker_recommendations(request):
    """AI-powered job recommendations for job seekers"""
    from applications.models import Application
    
    try:
        jobseeker = JobSeeker.objects.get(user=request.user)
    except JobSeeker.DoesNotExist:
        messages.error(request, "Job Seeker profile not found.")
        return redirect("jobseeker_dashboard")
    
    # Get jobs where user was hired - exclude these
    hired_job_ids = Application.objects.filter(
        applicant=jobseeker,
        status="Hired"
    ).values_list('job_id', flat=True)
    
    # Get all applications to track status
    applications = Application.objects.filter(applicant=jobseeker)
    applied_job_ids = [app.job_id for app in applications]
    applied_job_statuses = {app.job_id: app.status for app in applications}
    
    # ✅ CHECK IF USER HAS PREFERRED JOB TYPES (this is what we match on)
    has_profile_data = bool(jobseeker.preferred_job_types)
    
    # ✅ ONLY SHOW RECOMMENDATIONS IF USER HAS PREFERRED JOB TYPES
    recommendations = []
    
    if has_profile_data:
        # Get available jobs
        available_jobs = Job.objects.filter(
            is_available=True,
            is_approved=True
        ).exclude(
            employer__user=request.user
        ).exclude(
            id__in=hired_job_ids
        ).select_related('employer', 'employer__user')
        
        if available_jobs.exists():
            # ✅ SIMPLE MATCHING: Match preferred_job_types to job.category
            matched_jobs = []
            
            for job in available_jobs:
                # Calculate match score based on preferred job types
                score = 0
                
                if job.category:
                    # Direct category match
                    if job.category in jobseeker.preferred_job_types:
                        score = 100  # Perfect match
                    else:
                        # Partial match - check if any preferred type is similar
                        job_category_lower = job.category.lower()
                        for pref in jobseeker.preferred_job_types:
                            pref_lower = pref.lower()
                            # Check if words overlap
                            if pref_lower in job_category_lower or job_category_lower in pref_lower:
                                score = max(score, 70)  # Good match
                
                # Add to recommendations list
                matched_jobs.append((job, score))
            
            # Sort by score (highest first) and filter out 0% matches
            matched_jobs.sort(key=lambda x: x[1], reverse=True)
            recommendations = [(job, score) for job, score in matched_jobs if score > 0]
            
            # If no matches found, recommendations stays empty
    
    context = {
        "recommendations": recommendations,
        "applied_job_ids": applied_job_ids,
        "applied_job_statuses": applied_job_statuses,
        "seeker": jobseeker,
        "has_profile_data": has_profile_data,  # ✅ Based on preferred_job_types only
    }
    
    return render(request, "core/ai_job_recommendation.html", context)



# ====================================================
#                    NOTIFCATIONS
# ====================================================

@login_required
def get_notifications(request):
    notifications = Notification.objects.filter(receiver=request.user).order_by("-created_at")[:10]
    data = [
        {
            "message": n.message,
            "created_at": n.created_at.strftime("%b %d, %Y %I:%M %p"),
            "is_read": n.is_read,
        }
        for n in notifications
    ]
    return JsonResponse({"notifications": data})

@login_required
def mark_notifications_read(request):
    Notification.objects.filter(receiver=request.user, is_read=False).update(is_read=True)
    return JsonResponse({"status": "ok"})


from django.contrib.auth import get_user_model
User = get_user_model()
def login_view(request):
    return render(request, 'core/login.html')

from django.shortcuts import render, redirect
from django.contrib.auth import login
from django.contrib.auth import get_user_model
from .models import Employer, JobSeeker

User = get_user_model()

def social_login(request, provider):
    """
    🔹 Fake social login for BlueHire demo.
    Lets user choose role (Employer or JobSeeker) and employer type (Company or Personal).
    Creates distinct demo users and profiles automatically.
    """
    role = request.GET.get('role')               # 'employer' or 'jobseeker'
    emp_type = request.GET.get('type')           # 'company' or 'personal'

    # make unique identifier for each role/type combo
    identifier = f'{provider}_{role or "user"}_{emp_type or "default"}'
    email = f'demo_{identifier}@bluehire.com'

    # create or reuse demo user
    user, created = User.objects.get_or_create(
        username=email.split('@')[0],
        defaults={
            'email': email,
            'first_name': provider.capitalize()
        }
    )

    # handle employer or jobseeker
    if role == 'employer':
        Employer.objects.get_or_create(
            user=user,
            defaults={
                'company_name': 'Demo Company' if emp_type == 'company' else 'Personal Employer',
                'company_description': 'Demo employer profile created automatically.',
                'employer_type': emp_type or 'company'   # only if your model has this field
            }
        )
        redirect_url = '/system-admin/employer/dashboard/'

    elif role == 'jobseeker':
        JobSeeker.objects.get_or_create(user=user)
        redirect_url = '/system-admin/jobseeker/dashboard/'

    else:
        redirect_url = '/homepage/'

    # log the user in and redirect to proper dashboard
    login(request, user)
    return redirect(redirect_url)

def select_role(request):
    return render(request, 'core/select_role.html')

def select_employer_type(request):
    return render(request, 'core/select_employer_type.html')


# Replace your public_job_search function in accounts/views.py with this enhanced version

# Replace your public_job_search function in accounts/views.py

from django.core.paginator import Paginator
from django.db.models import Q, Count
from applications.models import Application

def public_job_search(request):
    """
    Enhanced public job search - Like Indeed
    Searches: title, description, category, company, location
    Filters: category, location, job_type (Full-time, Part-time, etc.)
    """
    
    # Get all search parameters
    query = request.GET.get('q', '').strip()
    location = request.GET.get('location', '').strip()
    category = request.GET.get('category', '').strip()
    job_type = request.GET.get('job_type', '').strip()  # NEW: Job Type filter
    sort_by = request.GET.get('sort', 'recent')
    
    print(f"[DEBUG] Search - Query: '{query}', Location: '{location}', Category: '{category}', Job Type: '{job_type}'")
    
    # Start with all approved and available jobs
    jobs = Job.objects.filter(
        is_approved=True,
        is_available=True
    ).select_related('employer', 'employer__user')
    
    # Apply search filters
    if query:
        jobs = jobs.filter(
            Q(title__icontains=query) |
            Q(description__icontains=query) |
            Q(category__icontains=query) |
            Q(employer__company_name__icontains=query)
        )
    
    if location:
        jobs = jobs.filter(location__icontains=location)
    
    if category:
        jobs = jobs.filter(category=category)
    
    # NEW: Job Type filter
    if job_type:
        jobs = jobs.filter(job_type__iexact=job_type)
    
    # Apply sorting
    if sort_by == 'title_asc':
        jobs = jobs.order_by('title')
    elif sort_by == 'title_desc':
        jobs = jobs.order_by('-title')
    elif sort_by == 'location':
        jobs = jobs.order_by('location', '-created_at')
    elif sort_by == 'company':
        jobs = jobs.order_by('employer__company_name', '-created_at')
    else:
        jobs = jobs.order_by('-created_at')
    
    # Get available categories
    categories = Job.objects.filter(
        is_approved=True,
        is_available=True
    ).values_list('category', flat=True).distinct().order_by('category')
    
    # Get available locations
    locations = Job.objects.filter(
        is_approved=True,
        is_available=True
    ).exclude(location__isnull=True).exclude(location='').values_list('location', flat=True).distinct().order_by('location')
    
    # Get available job types
    job_types = Job.objects.filter(
        is_approved=True,
        is_available=True
    ).exclude(job_type__isnull=True).exclude(job_type='').values_list('job_type', flat=True).distinct().order_by('job_type')
    
    # Get employers
    employers = Employer.objects.filter(
        jobs__is_approved=True,
        jobs__is_available=True
    ).distinct()
    
    # Check if user is logged in
    applied_job_ids = []
    applied_job_statuses = {}
    if request.user.is_authenticated and hasattr(request.user, 'jobseeker'):
        applications = Application.objects.filter(applicant=request.user.jobseeker)
        applied_job_ids = list(applications.values_list('job_id', flat=True))
        applied_job_statuses = {app.job_id: app.status for app in applications}
    
    # MANUALLY calculate stats for each job (THIS FIXES THE 0 APPLICANTS BUG)
    jobs_list = list(jobs)
    for job in jobs_list:
        # Count all applications for this job
        job.total_applicants = Application.objects.filter(job=job).count()
        job.hired_count = Application.objects.filter(job=job, status='Hired').count()
        job.rejected_count = Application.objects.filter(job=job, status='Rejected').count()
        job.pending_count = Application.objects.filter(job=job, status='Pending').count()
        
        print(f"[DEBUG] Job '{job.title}': {job.total_applicants} applicants, {job.hired_count} hired")
    
    context = {
        'jobs': jobs_list,
        'query': query,
        'location': location,
        'category': category,
        'job_type': job_type,
        'sort_by': sort_by,
        'total_results': len(jobs_list),
        'categories': categories,
        'locations': locations,
        'job_types': job_types,  # NEW
        'employers': employers,
        'applied_job_ids': applied_job_ids,
        'applied_job_statuses': applied_job_statuses,
    }
    
    return render(request, 'core/job_search_results.html', context)


# ====================================================
# 🔐 SOCIAL LOGIN HANDLERS (Google & Facebook)
# ====================================================
# ✅ ADD THESE THREE FUNCTIONS TO THE END OF accounts/views.py

# REPLACE the select_role function in accounts/views.py with this:

from allauth.socialaccount.models import SocialAccount

@login_required
def select_role(request):
    """
    After social login, user selects their role (JobSeeker or Employer)
    ✅ FIXED: Creates profiles with "Incomplete" status
    """
    user = request.user
    
    # ✅ CHECK: Only redirect if user has JobSeeker or Employer role
    if user.role in ['JobSeeker', 'Employer']:
        # User already has a valid role, redirect to dashboard
        if user.role == 'JobSeeker':
            return redirect('jobseeker_dashboard')
        elif user.role == 'Employer':
            return redirect('employer_dashboard')
    
    # ✅ If user is Admin OR has no role, show role selection
    if request.method == 'POST':
        role = request.POST.get('role')  # 'jobseeker' or 'employer'
        
        if role == 'jobseeker':
            # ✅ Create JobSeeker profile with "Incomplete" status
            seeker = JobSeeker.objects.create(user=user)
            seeker.approval_status = "Incomplete"
            seeker.verification_status = "Incomplete"
            seeker.save()
            
            user.role = 'JobSeeker'
            user.save()
            
            messages.success(request, "✅ Welcome to BlueHire! Please complete your job seeker profile.")
            return redirect('jobseeker_profile')
        
        elif role == 'employer':
            # Redirect to employer type selection
            return redirect('select_employer_type')
        
        else:
            messages.error(request, "Please select a valid role.")
            return redirect('select_role')
    
    return render(request, 'core/select_role.html', {'user': user})


@login_required
def select_employer_type(request):
    """
    Employer selects their type (Company or Personal)
    ✅ FIXED: Creates profiles with "Incomplete" status
    """
    user = request.user
    
    # Check if user already has employer profile
    try:
        employer = Employer.objects.get(user=user)
        if employer.employer_type:
            messages.info(request, "You already have an employer profile.")
            return redirect('employer_dashboard')
    except Employer.DoesNotExist:
        pass
    
    if request.method == 'POST':
        employer_type = request.POST.get('type')  # 'Company' or 'Personal'
        
        if employer_type in ['Company', 'Personal']:
            # ✅ Create Employer profile with "Incomplete" status
            employer, created = Employer.objects.get_or_create(
                user=user,
                defaults={
                    'employer_type': employer_type,
                    'approval_status': 'Incomplete',  # ✅ FIXED
                }
            )
            
            if not created:
                employer.employer_type = employer_type
                employer.approval_status = 'Incomplete'  # ✅ FIXED
                employer.save()
            
            user.role = 'Employer'
            user.save()
            
            messages.success(request, f"Welcome to BlueHire! Please complete your {employer_type} profile.")
            return redirect('employer_profile')
        
        else:
            messages.error(request, "Please select a valid employer type.")
            return redirect('select_employer_type')
    
    return render(request, 'core/select_employer_type.html', {'user': user})


@login_required
def complete_social_profile(request):
    """
    Optional: Let users add additional info after social login
    """
    user = request.user
    
    # Get social account info
    try:
        social_account = SocialAccount.objects.get(user=user)
        provider = social_account.provider  # 'google' or 'facebook'
        extra_data = social_account.extra_data
    except SocialAccount.DoesNotExist:
        provider = None
        extra_data = {}
    
    if request.method == 'POST':
        # Update user info
        user.phone = request.POST.get('phone', '')
        user.save()
        
        messages.success(request, "Profile completed successfully!")
        
        # Redirect based on role
        if user.role == 'JobSeeker':
            return redirect('jobseeker_dashboard')
        elif user.role == 'Employer':
            return redirect('employer_dashboard')
        else:
            return redirect('select_role')
    
    context = {
        'user': user,
        'provider': provider,
        'extra_data': extra_data,
    }
    
    return render(request, 'core/complete_social_profile.html', context)